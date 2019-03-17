from openpyxl import Workbook
import jieba.posseg as pseg
from openpyxl import load_workbook
import jieba
import os
import re
## 正则匹配文件是和哪个目标人物相关

owb = Workbook()  # 输出wb
ows = owb.active  # 获取活动表
ows.title = "Sheet1"  # 命名表
pattern = re.compile('{[^\.]+?[=]|,[^\.]+?[=]')
strinfo = re.compile(' ')
wb1 = load_workbook('./data/蒋劲夫_话题相关性标注-编辑.xlsx')  # 读入1

s_name1 = wb1.get_sheet_names()  # 得到输入表名数组

ws1 = wb1.get_sheet_by_name(s_name1[0])  # 得到读入book1的第一张表

print('sheet_name:%s' % s_name1[0])

f = open("data_new_j.txt", "w")

for i in ws1.row_dimensions:
    ##正则表达式修改得到target
    if i == 1 or i == 2:
        continue

    un_prefix = re.subn("#(.*)#",'',strinfo.sub('', ws1.cell(row=i, column=2).value))
    for w in pseg.cut(un_prefix[0]):
        f.write(w.word + " ")
    f.write('\n')
    f.write('蒋劲夫')
    f.write('\n')
    f.write(str(ws1.cell(row=i, column=1).value))
    f.write('\n')


f.close()